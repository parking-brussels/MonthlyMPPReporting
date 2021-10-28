# Class to read and write Excel files
import datetime
from openpyxl import Workbook, load_workbook
from office365.sharepoint.client_context import ClientContext

import pandas
from mpp import MPP as mppType


class ExcelFile:
    # First day of previous month
    # Calculated as beginning of this month minus 1 day, and take first day of that month)
    Month = (datetime.date.today().replace(day=1) - datetime.timedelta(days=1)).replace(day=1)

    # Open Excel
    def __init__(self, mpp):
        filename = mpp.GetFileName()
        self.wb = load_workbook(filename)

    # Add sheet for month (default last month)
    def AddWorkSheet(self, date=Month):
        self.wb.create_sheet(date.strftime("%mmm %Y"))

    # Add data to sheet and format sheet
    def AddData(self, pd):
        pass


# main script
if __name__ == '__main__':
    pass